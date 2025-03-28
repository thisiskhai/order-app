from collections import defaultdict
from datetime import timedelta, datetime
from django.utils.timezone import now
from django.utils.dateparse import parse_datetime
from django.contrib.auth import login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.models import User
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
import openpyxl

from .forms import ProductAdminForm, StaffCreateForm, ProductImportForm
from .models import Category, Product, OrderEntry

# ==========================
# ✅ User Permission Checks
# ==========================
def is_admin(user):
    return user.is_superuser

def is_staff_user(user):
    return user.is_staff and not user.is_superuser


# ==========================
# ✅ Homepage Redirect
# ==========================
@login_required
def product_list(request):
    if request.user.is_superuser:
        return redirect('admin_dashboard')
    elif request.user.is_staff:
        return redirect('staff_product_list')
    return redirect('staff_login')


# ==========================
# ✅ Login / Logout
# ==========================
def staff_login(request):
    form = AuthenticationForm(request, data=request.POST or None)
    if request.method == 'POST' and form.is_valid():
        user = form.get_user()
        if user.is_staff:
            login(request, user)
            next_url = request.GET.get('next')
            return redirect(next_url or ('admin_dashboard' if user.is_superuser else 'staff_product_list'))
        messages.error(request, "You do not have permission to access.")
    return render(request, 'staff_login.html', {'form': form})

@login_required
def staff_logout(request):
    logout(request)
    return redirect('staff_login')


# ==========================
# ✅ Admin Dashboard
# ==========================
@login_required
@user_passes_test(is_admin)
def admin_dashboard(request):
    staff_list = User.objects.filter(is_staff=True, is_superuser=False)
    staff_orders = [{
        'staff': staff,
        'total_quantity': OrderEntry.objects.filter(staff=staff).values('created_at').distinct().count()
    } for staff in staff_list]

    staff_form = StaffCreateForm(request.POST or None, prefix='staff')
    product_form = ProductAdminForm(request.POST or None, prefix='product')
    import_form = ProductImportForm(request.POST or None, request.FILES or None, prefix='import')

    if request.method == 'POST':
        if 'add_staff' in request.POST and staff_form.is_valid():
            staff_form.save()
            messages.success(request, "New staff added.")
            return redirect('admin_dashboard')
        if 'add_product' in request.POST and product_form.is_valid():
            product_form.save()
            messages.success(request, "New product added.")
            return redirect('admin_dashboard')
        if 'import_excel' in request.POST and import_form.is_valid():
            sheet = openpyxl.load_workbook(import_form.cleaned_data['excel_file']).active
            count = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                name, available, quantity, unit = row
                Product.objects.create(name=name, available_at_store=bool(available), order_quantity=int(quantity), unit=unit)
                count += 1
            messages.success(request, f"{count} products imported from Excel.")
            return redirect('admin_dashboard')

    products = Product.objects.all().order_by('-created_at')
    return render(request, 'admin_dashboard.html', {
        'staff_orders': staff_orders,
        'staff_form': staff_form,
        'product_form': product_form,
        'import_form': import_form,
        'products': products
    })


# ==========================
# ✅ Product Management
# ==========================
@login_required
@user_passes_test(is_admin)
def admin_product_list(request):
    products = Product.objects.all()
    form = ProductAdminForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Product added.")
        return redirect('admin_product_list')
    return render(request, 'admin_product_list.html', {'products': products, 'form': form})

@login_required
@user_passes_test(is_admin)
def product_update(request, uuid):
    product = get_object_or_404(Product, uuid=uuid)
    form = ProductAdminForm(request.POST or None, instance=product)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Product updated successfully!")
        return redirect('all_products')
    return render(request, 'product_form.html', {'form': form})

@login_required
@user_passes_test(is_admin)
def product_delete(request, uuid):
    product = get_object_or_404(Product, uuid=uuid)
    product.delete()
    messages.success(request, "Product deleted.")
    return redirect('admin_product_list')


# ==========================
# ✅ Product CRUD (Admin)
# ==========================
@login_required
@user_passes_test(is_admin)
def manage_products(request):
    products = Product.objects.all()
    product_form = ProductAdminForm(request.POST or None, prefix='product')
    import_form = ProductImportForm(request.POST or None, request.FILES or None, prefix='import')

    if request.method == 'POST':
        if 'add_product' in request.POST and product_form.is_valid():
            product_form.save()
            messages.success(request, "New product added.")
            return redirect('manage_products')
        if 'import_excel' in request.POST and import_form.is_valid():
            sheet = openpyxl.load_workbook(import_form.cleaned_data['excel_file']).active
            count = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                name, available, quantity, unit = row
                Product.objects.create(name=name, available_at_store=bool(available), order_quantity=int(quantity), unit=unit)
                count += 1
            messages.success(request, f"{count} products imported from Excel.")
            return redirect('manage_products')

    return render(request, 'manage_products.html', {
        'products': products,
        'product_form': product_form,
        'import_form': import_form,
    })

@login_required
@user_passes_test(is_admin)
def all_products(request):
    products = Product.objects.all().order_by('-created_at')
    return render(request, 'all_products.html', {'products': products})


# ==========================
# ✅ Staff Order Submission
# ==========================
@login_required
@user_passes_test(is_staff_user)
def staff_product_list(request):
    products = Product.objects.all()
    staff_orders = {e.product.id: e for e in OrderEntry.objects.filter(staff=request.user)}

    if request.method == 'POST':
        timestamp = now()
        new_entries = []

        for product in products:
            quantity = request.POST.get(f'quantity_{product.id}')
            if quantity:
                try:
                    quantity = int(quantity)
                    if quantity > 0:
                        new_entries.append(OrderEntry(
                            product=product, staff=request.user, order_quantity=quantity, created_at=timestamp
                        ))
                except ValueError:
                    messages.error(request, "Quantity must be a number.")

        if new_entries:
            old_ts = (
                OrderEntry.objects
                .filter(staff=request.user)
                .values_list('created_at', flat=True)
                .distinct()
                .order_by('created_at')
            )
            if len(old_ts) >= 5:
                OrderEntry.objects.filter(staff=request.user, created_at=old_ts[0]).delete()

            OrderEntry.objects.bulk_create(new_entries)
            messages.success(request, "Order submitted.")
            return redirect('staff_product_list')

    return render(request, 'staff_product_list.html', {
        'products': products,
        'staff_orders': staff_orders
    })

@login_required
@user_passes_test(is_staff_user)
def staff_order_history(request):
    staff = request.user
    products = Product.objects.all()
    timestamps = (
        OrderEntry.objects.filter(staff=staff)
        .values('created_at')
        .order_by('-created_at')
        .distinct()[:5]
    )
    orders_by_time = []
    for ts in timestamps:
        time = ts['created_at']
        entries = {e.product.id: e for e in OrderEntry.objects.filter(staff=staff, created_at=time)}
        orders_by_time.append({'timestamp': time, 'entries': entries})

    return render(request, 'staff_order_history.html', {
        'orders_by_time': orders_by_time,
        'products': products,
    })


# ==========================
# ✅ Admin: Orders per Staff
# ==========================
@login_required
@user_passes_test(is_admin)
def staff_order_detail(request, staff_id):
    staff = get_object_or_404(User, pk=staff_id, is_staff=True, is_superuser=False)

    timestamps = (
        OrderEntry.objects
        .filter(staff=staff)
        .values_list('created_at', flat=True)
        .distinct()
        .order_by('created_at')[:5]  # FIFO: oldest first
    )

    grouped_orders = []

    for ts in timestamps:
        entries = OrderEntry.objects.filter(staff=staff, created_at=ts).select_related('product')
        grouped_orders.append({
            'timestamp': ts,
            'entries': entries
        })

    return render(request, 'staff_order_detail.html', {
        'staff': staff,
        'grouped_orders': grouped_orders,
    })


# ==========================
# ✅ Admin: Order Batches
# ==========================
@login_required
@user_passes_test(is_admin)
def admin_order_batches(request, staff_id):
    staff = get_object_or_404(User, pk=staff_id, is_staff=True, is_superuser=False)
    timestamps = (
        OrderEntry.objects
        .filter(staff=staff)
        .order_by('-created_at')
        .values_list('created_at', flat=True)
        .distinct()
    )
    batches = []
    for ts in timestamps:
        entries = OrderEntry.objects.filter(staff=staff, created_at=ts).select_related('product')
        is_all_processed = all(entry.is_processed for entry in entries)
        batches.append({
            'timestamp': ts,
            'entries': entries,
            'is_all_processed': is_all_processed
        })

    return render(request, 'admin_order_batches.html', {
        'staff': staff,
        'batches': batches
    })

@login_required
@user_passes_test(is_admin)
def admin_batch_detail(request, staff_id, timestamp):
    staff = get_object_or_404(User, pk=staff_id, is_staff=True, is_superuser=False)

    try:
        created_at = datetime.strptime(timestamp, "%Y%m%d%H%M")
    except ValueError:
        return HttpResponse("❌ Invalid timestamp", status=400)

    entries = OrderEntry.objects.filter(
        staff=staff,
        created_at__year=created_at.year,
        created_at__month=created_at.month,
        created_at__day=created_at.day,
        created_at__hour=created_at.hour,
        created_at__minute=created_at.minute
    ).select_related('product')

    for entry in entries:
        if not entry.is_seen:
            entry.is_seen = True
            entry.save()

    if request.method == 'POST':
        for entry in entries:
            checkbox_name = f"processed_{entry.id}"
            is_checked = checkbox_name in request.POST
            if entry.is_processed != is_checked:
                entry.is_processed = is_checked
                entry.save()
        messages.success(request, "✅ Order status updated.")
        return redirect('admin_batch_detail', staff_id=staff.id, timestamp=timestamp)

    return render(request, 'admin_batch_detail.html', {
        'staff': staff,
        'timestamp': created_at,
        'entries': entries,
    })


# ==========================
# ✅ Other Admin Actions
# ==========================
@login_required
@user_passes_test(is_admin)
def toggle_order_status(request, order_id):
    order = get_object_or_404(OrderEntry, pk=order_id)
    order.is_processed = not order.is_processed
    order.save()
    return JsonResponse({'status': 'success', 'is_processed': order.is_processed})

@login_required
@user_passes_test(is_admin)
def delete_staff(request, staff_id):
    staff = get_object_or_404(User, pk=staff_id, is_staff=True, is_superuser=False)
    staff.delete()
    messages.success(request, f"Staff {staff.username} has been deleted.")
    return redirect('admin_dashboard')


# ==========================
# ✅ Cleanup Processed Orders older than 4 weeks
# ==========================
def cleanup_old_orders():
    OrderEntry.objects.filter(is_processed=True, created_at__lt=now() - timedelta(weeks=4)).delete()


@login_required
@user_passes_test(is_admin)
def all_products(request):
    categories = Category.objects.all()
    grouped_products = {
        category.name: Product.objects.filter(category=category).order_by('-created_at')
        for category in categories
    }

    return render(request, 'all_products.html', {
        'grouped_products': grouped_products
    })
